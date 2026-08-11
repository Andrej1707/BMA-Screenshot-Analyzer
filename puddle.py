import os
import re
import glob
import shutil
import sys
import threading
import queue
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict
import traceback
from datetime import datetime
import numpy as np
import openpyxl
import sys
import os

# Keep PaddleX in local-only mode before any Paddle imports run.
os.environ.setdefault("PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK", "True")

def _patch_paddlex_deps_check_for_exe():
    """
    PaddleOCR>=3 uses PaddleX pipeline.
    In PyInstaller onefile/onedir, PaddleX may not see package metadata and falsely
    raises: DependencyError: `OCR` requires additional dependencies.
    We bypass ONLY in frozen exe. If deps are truly missing, it will fail later anyway.
    """
    if not getattr(sys, "frozen", False):
        return

    try:
        import paddlex.utils.deps as deps
        deps.require_extra = lambda *args, **kwargs: None
    except Exception:
        # if paddlex isn't importable, nothing we can do here
        pass

_patch_paddlex_deps_check_for_exe()
from paddleocr import PaddleOCR

# --- PyInstaller keepalive imports (wichtig für PaddleX Pipeline Registry) ---
import paddlex  # noqa: F401
import paddlex.inference.pipelines.ocr.pipeline  # noqa: F401
import paddlex.inference.pipelines._parallel  # noqa: F401

def normalize_element_typ(s: str) -> str:
    if not s:
        return s
    s = s.strip()

    # 0/O Fix nur im Element-Typ Kontext (FAP-...)
    # D0 -> DO, D0T -> DOT
    s = re.sub(r"^(FAP-)D0T(\d+)$", r"\1DOT\2", s)  # FAP-D0T425 -> FAP-DOT425
    s = re.sub(r"^(FAP-)D0(\d+)$",  r"\1DO\2",  s)  # FAP-D0425  -> FAP-DO425
    s = re.sub(r"^(FAP-)0(\d+)$", r"\1O\2", s)      # FAP-0425 -> FAP-O425
    s = re.sub(r"^(FAP-)o(\d+)$", r"\1O\2", s)      # FAP-o425 -> FAP-O425
    s = re.sub(r"[^A-Za-z0-9\-]", "", s)

    return s

# =========================
# APP / BRANDING
# =========================
APP_TITLE = "BMA Screenshot Analyzer"
APP_SUBTITLE = "Diagnose-Screenshots zu Ring-Excel, vollstaendig lokal"


def resource_path(relative_path: str) -> str:
    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(os.path.abspath(sys.executable))
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, relative_path)


# =========================
# EXCEL MAPPING
# =========================
SHEET_NAME = "Tabelle1"

SITE_CELL_ROW, SITE_CELL_COL = 3, 2   # B3
RING_CELL_ROW, RING_CELL_COL = 5, 2   # B5
LSN_CELL_ROW,  LSN_CELL_COL  = 6, 2   # B6

DATA_START_ROW = 8
CLEAR_UNTIL_ROW = 250

DATA_COLS = {
    "Nr": 3,
    "Element": 4,
    "Adresse": 5,
    "Btr": 6,
    "CO": 7,
    "Vrs": 8,
}

EXPECTED_ROWS_PER_IMAGE = 7


# =========================
# UTIL
# =========================
def ring_folder_plain(n: int) -> str:
    return f"Ring {n}"

def ring_folder_named(n: int, lsn: int) -> str:
    return f"Ring {n} LSN {lsn}"

def parse_ring_num(name: str) -> Optional[int]:
    m = re.search(r"ring\s*(\d+)", name, re.IGNORECASE)
    return int(m.group(1)) if m else None

def parse_lsn_num(name: str) -> Optional[int]:
    m = re.search(r"lsn\s*(\d+)", name, re.IGNORECASE)
    return int(m.group(1)) if m else None


# =========================
# OCR NORMALIZE (ROBUST)
# =========================
HEX_RE = re.compile(r"^0x[0-9A-Fa-f]+$")
ADDR_RE = re.compile(r"^\d{1,3}-\d{1,2}$")
BTR_RE  = re.compile(r"^\d{1,6}$")
ELEM_RE = re.compile(r"^[A-Za-z0-9\-]+$")

def normalize_hex(s: str) -> str:
    """
    OCR-proof hex normalize:
    - O/o -> 0
    - l/I/|/! -> 1
    - strip garbage chars
    - ensure 0x prefix
    """
    s = (s or "").strip()

    s = s.replace("Ox", "0x").replace("ox", "0x").replace("0X", "0x")
    s = s.replace("×", "x")

    # common OCR swaps
    s = s.replace("O", "0").replace("o", "0")
    s = s.replace("l", "1").replace("I", "1").replace("|", "1").replace("!", "1")

    # remove junk
    s = re.sub(r"[^0-9a-fA-Fx]", "", s)

    if s.startswith("x"):
        s = "0" + s

    # if "0A" like -> force 0x
    if re.fullmatch(r"[0-9A-Fa-f]+", s):
        s = "0x" + s
    elif re.fullmatch(r"0[0-9A-Fa-f]+", s) and not s.startswith("0x"):
        s = "0x" + s[1:]

    return s.lower()

def normalize_nr(s: str) -> str:
    """
    IMPORTANT FIX (Paddle relevant):
    Old Tesseract-world logic would extract digits from 'FAP-DOT425' => '425' and kill element parsing.
    Now: Nr is only accepted if token has NO letters at all.
    """
    s = (s or "").strip()
    if not s:
        return ""

    # If the token contains letters, it is NOT a pure Nr field.
    if re.search(r"[A-Za-z]", s):
        return ""

    s = s.replace("O", "0").replace("o", "0")
    s = s.replace("B", "8")
    s = s.replace("S", "5")
    s = re.sub(r"[^0-9]", "", s)
    return s

def normalize_addr(tok: str) -> str:
    tok = (tok or "").strip()
    tok = tok.replace("_", "-").replace("—", "-").replace("–", "-").replace(" ", "")
    # "455" -> "45-5"
    if re.fullmatch(r"\d{3}", tok):
        return tok[:2] + "-" + tok[2]
    return tok

def merge_addr_tokens(tokens: List[str]) -> List[str]:
    """
    Merge addr patterns split by OCR: ["45","-","5"] -> "45-5"
    """
    out = []
    i = 0
    while i < len(tokens):
        if i + 2 < len(tokens) and tokens[i].isdigit() and tokens[i+1] in ["-", "–", "—"] and tokens[i+2].isdigit():
            out.append(tokens[i] + "-" + tokens[i+2])
            i += 3
            continue
        if i + 1 < len(tokens) and tokens[i].endswith("-") and tokens[i][:-1].isdigit() and tokens[i+1].isdigit():
            out.append(tokens[i][:-1] + "-" + tokens[i+1])
            i += 2
            continue
        out.append(tokens[i])
        i += 1
    return out


# =========================
# PARSING ROWS (FLEXIBLE)
# =========================
@dataclass
class Row:
    nr: int
    element: str
    adresse: str
    btr: int
    co: str
    vrs: str
    conf: float = 1.0

def parse_row_generic(line: str) -> Optional[dict]:
    """
    Parse row line into fields:
    Expect (roughly): Nr Element Adresse Btr CO Vrs
    But Nr can be missing -> we allow nr=None and fill later.
    """
    toks = [t for t in re.split(r"\s+", (line or "").strip()) if t]
    if len(toks) < 5:
        return None

    toks = merge_addr_tokens(toks)

    co = normalize_hex(toks[-2])
    vrs = normalize_hex(toks[-1])
    if not HEX_RE.fullmatch(co) or not HEX_RE.fullmatch(vrs):
        return None

    btr_tok = toks[-3].replace("O", "0").replace("o", "0")
    if not BTR_RE.fullmatch(btr_tok):
        return None
    btr = int(btr_tok)

    addr_tok = normalize_addr(toks[-4])
    if not ADDR_RE.fullmatch(addr_tok):
        return None

    prefix = toks[:-4]
    if not prefix:
        return None

    nr_val = None
    nr_s = normalize_nr(prefix[0])
    if nr_s:
        nr_val = int(nr_s)
        elem_parts = prefix[1:]
    else:
        elem_parts = prefix

    element = "".join(elem_parts).replace(" ", "")
    element = normalize_element_typ(element)
    if not element or not ELEM_RE.fullmatch(element):
        return None

    return {
        "nr": nr_val,
        "element": element,
        "adresse": addr_tok,
        "btr": btr,
        "co": co,
        "vrs": vrs
    }

def finalize_nr_sequence(rows_with_y: List[dict]) -> List[Row]:
    """
    Fill missing nr based on y-order. Also fix decreasing/duplicate nr.
    Dedupe by (nr) at the end.
    """
    if not rows_with_y:
        return []

    rows_with_y.sort(key=lambda x: x["y"])

    prev = None
    for r in rows_with_y:
        if r["nr"] is None:
            r["nr"] = 1 if prev is None else prev + 1
        else:
            if prev is not None and r["nr"] <= prev:
                r["nr"] = prev + 1
        prev = r["nr"]

    out = [Row(r["nr"], r["element"], r["adresse"], r["btr"], r["co"], r["vrs"], float(r.get("conf", 1.0)))
       for r in rows_with_y]

    # dedupe by nr (overlap screenshots)
    ded = {}
    for rr in out:
        ded[rr.nr] = rr
    final = list(ded.values())
    final.sort(key=lambda x: x.nr)
    return final


# =========================
# PADDLE OCR (LOCAL MODELS, SINGLETON)
# =========================
_OCR_SINGLETON: Optional[PaddleOCR] = None
_OCR_KEY: Optional[Tuple[str, str]] = None  # (det_dir, rec_dir)

def get_paddle_ocr(project_dir: str) -> PaddleOCR:
    """
    Creates or returns a cached PaddleOCR instance using local model dirs.
    Expected model layout:
      project_dir\\OCR-Modelle\\PP-OCRv5_server_det
      project_dir\\OCR-Modelle\\PP-OCRv5_server_rec
    """
    global _OCR_SINGLETON, _OCR_KEY

    models_dir = os.path.join(project_dir, "OCR-Modelle")
    det_dir = os.path.join(models_dir, "PP-OCRv5_server_det")
    rec_dir = os.path.join(models_dir, "PP-OCRv5_server_rec")

    if not os.path.isdir(det_dir):
        raise FileNotFoundError(f"❌ Missing det model dir: {det_dir}")
    if not os.path.isdir(rec_dir):
        raise FileNotFoundError(f"❌ Missing rec model dir: {rec_dir}")

    key = (det_dir, rec_dir)
    if _OCR_SINGLETON is not None and _OCR_KEY == key:
        return _OCR_SINGLETON

    # Disable remote hoster checks (you already used this successfully)
    os.environ.setdefault("PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK", "True")

    # IMPORTANT:
    # - lang is ignored when model dirs are set (fine)
    try:
        ocr = PaddleOCR(
            text_detection_model_dir=det_dir,
            text_recognition_model_dir=rec_dir,

            # WICHTIG: verhindert Download zusätzlicher Modelle
            use_doc_orientation_classify=False,
            use_doc_unwarping=False,
            use_textline_orientation=False,

        )
    except Exception:
        tb = traceback.format_exc()
        log_path = os.path.join(project_dir, "paddle_error.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n" + "="*80 + "\n")
            f.write(f"{datetime.now().isoformat()}  PaddleOCR init failed\n")
            f.write(tb)
        raise

    _OCR_SINGLETON = ocr
    _OCR_KEY = key
    return ocr


def paddle_tokens_to_lines(result_item: dict) -> List[Tuple[float, str, float]]:
    """
    Convert Paddle output into ordered text lines:
    Returns list[(y, "joined tokens", line_conf)]
    line_conf = mean(rec_scores of tokens in that line)
    """
    texts = result_item.get("rec_texts") or []
    polys = result_item.get("rec_polys") or []
    scores = result_item.get("rec_scores") or []

    tokens = []
    for i, (t, p) in enumerate(zip(texts, polys)):
        if not t:
            continue
        p = np.asarray(p)
        x_min = float(np.min(p[:, 0]))
        y_min = float(np.min(p[:, 1]))
        y_max = float(np.max(p[:, 1]))
        y_center = (y_min + y_max) / 2.0

        sc = float(scores[i]) if i < len(scores) else 1.0
        tokens.append((y_center, x_min, str(t).strip(), sc))

    if not tokens:
        return []

    tokens.sort(key=lambda z: (z[0], z[1]))

    ys = [t[0] for t in tokens]
    deltas = [ys[i+1] - ys[i] for i in range(len(ys)-1)]
    pos = [d for d in deltas if d > 0.5]
    thr = float(np.median(pos)) if pos else 14.0
    thr = max(8.0, min(thr, 22.0))

    lines: List[List[Tuple[float, float, str, float]]] = []
    cur = [tokens[0]]
    cur_y = tokens[0][0]

    for tok in tokens[1:]:
        y = tok[0]
        if abs(y - cur_y) <= thr:
            cur.append(tok)
            cur_y = (cur_y * 0.7) + (y * 0.3)
        else:
            lines.append(cur)
            cur = [tok]
            cur_y = y
    lines.append(cur)

    out = []
    for line in lines:
        line.sort(key=lambda z: z[1])  # x
        y_avg = float(np.mean([z[0] for z in line]))
        txt = " ".join([z[2] for z in line]).strip()
        conf = float(np.mean([z[3] for z in line])) if line else 1.0
        out.append((y_avg, txt, conf))

    out.sort(key=lambda t: t[0])
    return out


# =========================
# IMAGE -> ROWS (PADDLE)
# =========================
def extract_rows_from_image(img_path: str, project_dir: str) -> Tuple[List[Row], List[Tuple[float, str]], int]:
    """
    Paddle-based extraction:
    - OCR the image (full image)
    - build lines by y clustering
    - parse rows via parse_row_generic
    Returns:
      (rows, debug_lines, token_count)
    """
    ocr = get_paddle_ocr(project_dir)
    result = ocr.predict(img_path)

    if not result or not isinstance(result, list) or not result[0]:
        return ([], [], 0)

    item = result[0]
    lines = paddle_tokens_to_lines(item)
    token_count = len(item.get("rec_texts") or [])

    # parse candidates
    cands = []
    for y, ln, conf in lines:
        r = parse_row_generic(ln)
        if r:
            r["y"] = y
            r["conf"] = conf
            cands.append(r)

    rows = finalize_nr_sequence(cands)

    # keep only first EXPECTED rows per screenshot (same behavior as old tool)
    if len(rows) > EXPECTED_ROWS_PER_IMAGE:
        rows = rows[:EXPECTED_ROWS_PER_IMAGE]

    return (rows, lines, token_count)


# =========================
# EXCEL OUTPUT (STYLE SAFE)
# =========================
def write_excel(template_path: str, out_path: str, site_name: str, ring_num: int, lsn_num: int, rows: List[Row]):
    if os.path.exists(out_path):
        os.remove(out_path)
    shutil.copy(template_path, out_path)

    wb = openpyxl.load_workbook(out_path)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"Sheet '{SHEET_NAME}' nicht gefunden.")
    ws = wb[SHEET_NAME]

    # Header
    ws.cell(row=SITE_CELL_ROW, column=SITE_CELL_COL).value = site_name
    ws.cell(row=RING_CELL_ROW, column=RING_CELL_COL).value = ring_num
    ws.cell(row=LSN_CELL_ROW,  column=LSN_CELL_COL).value = lsn_num

    # Reference styles from first data row (template)
    ref_styles = {}
    for _, col in DATA_COLS.items():
        ref_styles[col] = ws.cell(row=DATA_START_ROW, column=col)._style

    # Clear area and enforce style
    for r in range(DATA_START_ROW, CLEAR_UNTIL_ROW):
        for _, col in DATA_COLS.items():
            c = ws.cell(row=r, column=col)
            c.value = None
            c._style = ref_styles[col]

    # Write data and enforce style for each cell (prevents random bold)
    for i, row in enumerate(rows):
        rr = DATA_START_ROW + i

        ws.cell(row=rr, column=DATA_COLS["Nr"]).value = row.nr
        ws.cell(row=rr, column=DATA_COLS["Element"]).value = row.element
        ws.cell(row=rr, column=DATA_COLS["Adresse"]).value = row.adresse
        ws.cell(row=rr, column=DATA_COLS["Btr"]).value = row.btr
        ws.cell(row=rr, column=DATA_COLS["CO"]).value = row.co
        ws.cell(row=rr, column=DATA_COLS["Vrs"]).value = row.vrs

        for _, col in DATA_COLS.items():
            ws.cell(row=rr, column=col)._style = ref_styles[col]

    wb.save(out_path)


# =========================
# RUN PER RING
# =========================
@dataclass
class RingCfg:
    ring_num: int
    lsn_num: int
    folder: str

def run_project(project_dir: str, site_name: str, rings: List[RingCfg], log):
    template = os.path.join(project_dir, "MelderService Grundexcel.xlsx")
    if not os.path.exists(template):
        raise FileNotFoundError("❌ 'MelderService Grundexcel.xlsx' fehlt im Projektordner.")

    # Paddle model dirs check early
    _ = get_paddle_ocr(project_dir)

    for rcfg in rings:
        ring_dir = rcfg.folder

        imgs = []
        for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
            imgs.extend(glob.glob(os.path.join(ring_dir, ext)))
        imgs.sort()

        if not imgs:
            log(f"⚠️ {os.path.basename(ring_dir)}: keine Bilder → übersprungen.\n")
            continue

        all_rows: List[Row] = []
        for ip in imgs:
            try:
                rows, _lines, tok_count = extract_rows_from_image(ip, project_dir=project_dir)

                # Do not expose OCR-derived screenshot contents in logs.
                log(f"🖼 {os.path.basename(ip)} | TOKENS: {tok_count} | ROWS: {len(rows)}\n")

                all_rows.extend(rows)

            except Exception as e:
                log(f"⚠️ Bild übersprungen (Fehler): {os.path.basename(ip)} → {e}\n")
                continue
        # ✅ Dedupe by Adresse: keep ONLY one row per address.
        # If address appears multiple times (overlap screenshots), keep higher confidence.
        best_by_addr: Dict[str, Row] = {}
        order: List[str] = []

        for r in all_rows:
            a = r.adresse
            if a not in best_by_addr:
                best_by_addr[a] = r
                order.append(a)
            else:
                if r.conf > best_by_addr[a].conf:
                    best_by_addr[a] = r

        merged = [best_by_addr[a] for a in order]

        # ✅ GLOBAL renumbering AFTER merge (fixes per-screenshot reset / missing Nrs)
        for idx, r in enumerate(merged, start=1):
            r.nr = idx

        out_path = os.path.join(ring_dir, f"MelderService Ring {rcfg.ring_num}.xlsx")
        write_excel(template, out_path, site_name, rcfg.ring_num, rcfg.lsn_num, merged)

        log(f"✅ {os.path.basename(ring_dir)}: {len(imgs)} Bilder → {len(merged)} Zeilen → {os.path.basename(out_path)}\n\n")


# =========================
# GUI
# =========================
class GuiLogger:
    def __init__(self, q: queue.Queue):
        self.q = q
    def write(self, s):
        if s:
            self.q.put(s)
    def flush(self):
        pass


def start_gui():
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog

    qlog = queue.Queue()

    def log(s: str):
        qlog.put(s)

    root = tk.Tk()
    root.title(APP_TITLE)
    root.geometry("980x680")
    try:
        root.iconbitmap(resource_path(os.path.join("Dependencies", "assets", "app.ico")))
    except Exception:
        pass

    # =========================
    # BUSY OVERLAY (Spinner + Input Block)
    # =========================
    class BusyOverlay:
        def __init__(self, root_):
            self.root = root_
            self.win = None
            self._job = None
            self._angle = 0

        def start(self, text="Bitte warten…", topmost=True):
            if self.win and self.win.winfo_exists():
                try:
                    self.label.config(text=text)
                    if topmost:
                        self.win.attributes("-topmost", True)
                    self.win.deiconify()
                    self.win.lift()
                    self.win.focus_force()
                    self.win.update_idletasks()
                    return
                except Exception:
                    pass

            self.win = tk.Toplevel(self.root)
            self.win.withdraw()
            self.win.overrideredirect(True)  # kein Rahmen
            self.win.attributes("-topmost", bool(topmost))
            self.win.configure(bg="#f3f3f3")

            # zentrieren (auf Bildschirm)
            w, h = 220, 120
            sw = self.win.winfo_screenwidth()
            sh = self.win.winfo_screenheight()
            x = (sw - w) // 2
            y = (sh - h) // 2
            self.win.geometry(f"{w}x{h}+{x}+{y}")

            card = tk.Frame(self.win, bg="#f3f3f3")
            card.pack(expand=True, fill="both")

            self.canvas = tk.Canvas(card, width=44, height=44, bg="#f3f3f3", highlightthickness=0)
            self.canvas.pack(pady=(18, 6))

            r = 18
            cx, cy = 22, 22
            self.canvas.create_oval(cx-r, cy-r, cx+r, cy+r, outline="#d0d0d0", width=4)
            self.arc = self.canvas.create_arc(cx-r, cy-r, cx+r, cy+r, start=0, extent=60,
                                              style="arc", outline="#7a7a7a", width=4)

            self.label = tk.Label(card, text=text, bg="#f3f3f3", fg="#555", font=("Segoe UI", 10))
            self.label.pack()

            self.win.deiconify()
            self.win.lift()
            self.win.update_idletasks()

            # blockt Eingaben global
            try:
                self.win.grab_set_global()
            except Exception:
                try:
                    self.win.grab_set()
                except Exception:
                    pass

            self._spin()

        def _spin(self):
            self._angle = (self._angle + 15) % 360
            try:
                self.canvas.itemconfig(self.arc, start=self._angle)
            except Exception:
                return
            self._job = self.root.after(30, self._spin)

        def stop(self):
            if self._job:
                try:
                    self.root.after_cancel(self._job)
                except Exception:
                    pass
                self._job = None

            if self.win and self.win.winfo_exists():
                try:
                    self.win.grab_release()
                except Exception:
                    pass
                try:
                    self.win.withdraw()
                except Exception:
                    pass

    root.withdraw()              # Hauptfenster weg
    busy = BusyOverlay(root)     # jetzt existiert die Klasse, safe
    busy.start("Starte…", topmost=True)
    root.update_idletasks()

    def run_busy(fn, text="Bitte warten…"):
        busy.start(text)
        root.update_idletasks()
        try:
            return fn()
        finally:
            busy.stop()

    def run_busy_async(fn, on_ok=None, on_err=None, text="Bitte warten…"):
        """
        fn läuft im Thread (kein Tk inside!).
        callbacks laufen im Tk thread.
        """
        busy.start(text)
        root.update_idletasks()

        def bg():
            try:
                res = fn()
            except Exception as e:
                if on_err:
                    root.after(0, lambda: (busy.stop(), on_err(e)))
                else:
                    root.after(0, lambda: (busy.stop(), messagebox.showerror("Fehler", str(e))))
            else:
                if on_ok:
                    root.after(0, lambda: (busy.stop(), on_ok(res)))
                else:
                    root.after(0, busy.stop)

        threading.Thread(target=bg, daemon=True).start()


    header = tk.Frame(root)
    header.pack(fill="x", padx=12, pady=10)

    tf = tk.Frame(header)
    tf.pack(fill="x", expand=True)
    tk.Label(tf, text=APP_TITLE, font=("Segoe UI", 16, "bold")).pack(anchor="w")
    tk.Label(tf, text=APP_SUBTITLE, font=("Segoe UI", 10)).pack(anchor="w")

    container = tk.Frame(root)
    container.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    log_frame = ttk.LabelFrame(root, text="Log")
    log_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    text = tk.Text(log_frame, wrap="word")
    text.pack(fill="both", expand=True, padx=10, pady=10)
    text.insert("end", f"{APP_TITLE}\n\n")

    def pump_queue():
        try:
            while True:
                s = qlog.get_nowait()
                text.insert("end", s)
                text.see("end")
        except queue.Empty:
            pass
        root.after(80, pump_queue)

    pump_queue()

    state = {
        "project_dir": "",
        "site": "",
        "ring_count": 0,
        "ring_folders": {},
        "lsn_map": {},
        "rings_final": [],
        "running": False,
    }

    def show(frame):
        for child in container.winfo_children():
            child.pack_forget()
        frame.pack(fill="both", expand=True)

    # STEP 1
    step1 = ttk.LabelFrame(container, text="Schritt 1/4 – Projektordner & Standort")
    step1_inner = tk.Frame(step1)
    step1_inner.pack(fill="x", padx=10, pady=10)

    proj_var = tk.StringVar(value="")
    site_var = tk.StringVar(value="")

    def suggest_default_dir():
        if getattr(sys, "frozen", False):
            return os.path.dirname(os.path.abspath(sys.executable))
        return os.path.dirname(os.path.abspath(__file__))

    proj_var.set(suggest_default_dir())

    tk.Label(step1_inner, text="Arbeitsordner (EXE/Grundexcel/OCR-Modelle/Ring-Ordner):").pack(anchor="w")
    proj_entry = ttk.Entry(step1_inner, textvariable=proj_var)
    proj_entry.pack(fill="x", pady=6)

    def browse_project():
        d = filedialog.askdirectory(title="Projektordner wählen", initialdir=proj_var.get() or suggest_default_dir())
        if d:
            proj_var.set(d)

    ttk.Button(step1_inner, text="Ordner auswählen...", command=browse_project).pack(anchor="w")

    tk.Label(step1_inner, text="Standort (wird in Excel nach B3 geschrieben):").pack(anchor="w", pady=(14, 0))
    site_entry = ttk.Entry(step1_inner, textvariable=site_var)
    site_entry.pack(fill="x", pady=6)

    def validate_project_dir(d: str) -> Tuple[bool, str]:
        if not d or not os.path.isdir(d):
            return False, "Bitte einen gültigen Ordner auswählen."
        if not os.path.exists(os.path.join(d, "MelderService Grundexcel.xlsx")):
            return False, "Fehlt im Projektordner: MelderService Grundexcel.xlsx"
        # Models must exist
        det_dir = os.path.join(d, "OCR-Modelle", "PP-OCRv5_server_det")
        rec_dir = os.path.join(d, "OCR-Modelle", "PP-OCRv5_server_rec")
        if not os.path.isdir(det_dir):
            return False, f"Fehlt: OCR-Modelle\\PP-OCRv5_server_det\n({det_dir})"
        if not os.path.isdir(rec_dir):
            return False, f"Fehlt: OCR-Modelle\\PP-OCRv5_server_rec\n({rec_dir})"
        return True, ""

    nav1 = tk.Frame(step1)
    nav1.pack(fill="x", padx=10, pady=10)

    def step1_next():
        d = proj_var.get().strip()
        site = site_var.get().strip()

        ok, err = validate_project_dir(d)
        if not ok:
            messagebox.showerror("Projektordner nicht bereit", err)
            return
        if not site:
            messagebox.showerror("Standort fehlt", "Bitte Standort eintragen (Excel B3).")
            return

        state["project_dir"] = d
        state["site"] = site

        log(f"📌 Projektordner gesetzt: {d}\n")
        log(f"🏷 Standort gesetzt (B3): {site}\n")

        def do_warmup():
            _ = get_paddle_ocr(d)  # darf dauern
            return True

        def ok_cb(_):
            log("🧠 PaddleOCR bereit (lokale Models).\n\n")
            show(step2)

        def err_cb(e):
            messagebox.showerror("PaddleOCR Fehler", str(e))

        run_busy_async(do_warmup, on_ok=ok_cb, on_err=err_cb, text="Lade PaddleOCR…")

    ttk.Button(nav1, text="Weiter →", command=step1_next).pack(side="right")

    # STEP 2
    step2 = ttk.LabelFrame(container, text="Schritt 2/4 – Ring-Ordner erstellen")
    s2 = tk.Frame(step2)
    s2.pack(fill="x", padx=10, pady=10)

    tk.Label(s2, text="Wie viele Ringe brauchst du?", font=("Segoe UI", 10, "bold")).pack(anchor="w")
    ring_count_var = tk.IntVar(value=6)
    spin = ttk.Spinbox(s2, from_=1, to=50, textvariable=ring_count_var, width=8)
    spin.pack(anchor="w", pady=6)

    explain = tk.Label(
        s2,
        text="Es werden nur 'Ring 1', 'Ring 2', ... erstellt.\nLSN kommt im nächsten Schritt (Umbenennen).",
        foreground="#444"
    )
    explain.pack(anchor="w", pady=(8, 0))

    nav2 = tk.Frame(step2)
    nav2.pack(fill="x", padx=10, pady=10)

    def step2_back():
        run_busy(lambda: show(step1), "Wechsle Schritt…")

    def step2_next_create():
        def work():
            project_dir = state["project_dir"]
            n = int(ring_count_var.get())
            state["ring_count"] = n
            state["ring_folders"].clear()

            created = 0
            for i in range(1, n + 1):
                p = os.path.join(project_dir, ring_folder_plain(i))
                if not os.path.isdir(p):
                    os.makedirs(p, exist_ok=True)
                    created += 1
                state["ring_folders"][i] = p

            log(f"📁 Ring-Ordner bereit: {n} (neu erstellt: {created})\n")

            try:
                if os.name == "nt":
                    os.startfile(project_dir)
            except Exception:
                pass

            build_step3_rows()
            show(step3)

        run_busy(work, "Erstelle Ring-Ordner…")

    ttk.Button(nav2, text="← Zurück", command=step2_back).pack(side="left")
    ttk.Button(nav2, text="Ordner erstellen & weiter →", command=step2_next_create).pack(side="right")

    # STEP 3
    step3 = ttk.LabelFrame(container, text="Schritt 3/4 – LSN eintragen & Ordner umbenennen")
    s3 = tk.Frame(step3)
    s3.pack(fill="both", expand=True, padx=10, pady=10)

    tk.Label(s3, text="Trag pro Ring die Modulnummer (LSN) ein. Dann 'Speichern & Umbenennen'.", foreground="#444").pack(anchor="w")

    table = tk.Frame(s3)
    table.pack(fill="both", expand=True, pady=10)

    rows_widgets = {}

    def build_step3_rows():
        for w in table.winfo_children():
            w.destroy()
        rows_widgets.clear()
        state["lsn_map"].clear()

        header = tk.Frame(table)
        header.pack(fill="x", pady=(0, 6))
        tk.Label(header, text="Ring", width=10, anchor="w", font=("Segoe UI", 9, "bold")).pack(side="left")
        tk.Label(header, text="LSN", width=12, anchor="w", font=("Segoe UI", 9, "bold")).pack(side="left")
        tk.Label(header, text="Ordnername (wird so umbenannt)", anchor="w", font=("Segoe UI", 9, "bold")).pack(side="left", fill="x", expand=True)

        for ring in range(1, state["ring_count"] + 1):
            r = tk.Frame(table)
            r.pack(fill="x", pady=2)

            tk.Label(r, text=f"{ring}", width=10, anchor="w").pack(side="left")

            lsn_var = tk.StringVar(value="")
            ent = ttk.Entry(r, textvariable=lsn_var, width=10)
            ent.pack(side="left", padx=(0, 12))

            current_name = os.path.basename(state["ring_folders"][ring])
            existing_lsn = parse_lsn_num(current_name)
            if existing_lsn is not None:
                lsn_var.set(str(existing_lsn))
                ent.state(["disabled"])

            folder_lbl = tk.Label(r, text=f"→ (LSN fehlt/ungültig)", anchor="w")
            folder_lbl.pack(side="left", fill="x", expand=True)

            def make_update(lbl, rv, ringnum):
                def _upd(*_):
                    v = rv.get().strip()
                    if v.isdigit():
                        lbl.config(text=f"→ {ring_folder_named(ringnum, int(v))}")
                    else:
                        lbl.config(text="→ (LSN fehlt/ungültig)")
                return _upd

            lsn_var.trace_add("write", make_update(folder_lbl, lsn_var, ring))
            make_update(folder_lbl, lsn_var, ring)()

            rows_widgets[ring] = (lsn_var, folder_lbl)

    def step3_back():
        run_busy(lambda: show(step2), "Wechsle Schritt…")

    def step3_save_and_rename():
        project_dir = state["project_dir"]

        lsns = {}
        for ring in range(1, state["ring_count"] + 1):
            lsn_var, _ = rows_widgets[ring]
            v = lsn_var.get().strip()
            if not v.isdigit():
                cur_name = os.path.basename(state["ring_folders"][ring])
                ex = parse_lsn_num(cur_name)
                if ex is None:
                    raise ValueError(f"Ring {ring}: LSN fehlt oder ist ungültig.")
                lsns[ring] = ex
                continue
            lsns[ring] = int(v)

        for ring, lsn in lsns.items():
            current_path = state["ring_folders"][ring]
            cur_base = os.path.basename(current_path)

            if re.match(rf"^Ring\s*{ring}\s+LSN\s+{lsn}\b", cur_base, re.IGNORECASE):
                continue

            target = os.path.join(project_dir, ring_folder_named(ring, lsn))
            if os.path.abspath(current_path) != os.path.abspath(target):
                if os.path.exists(target):
                    raise FileExistsError(f"Zielordner existiert schon: {os.path.basename(target)}")
                os.rename(current_path, target)
                state["ring_folders"][ring] = target

        state["lsn_map"] = lsns

        final = []
        for ring in range(1, state["ring_count"] + 1):
            final.append(RingCfg(ring_num=ring, lsn_num=state["lsn_map"][ring], folder=state["ring_folders"][ring]))
        state["rings_final"] = final

        log("✅ LSN gespeichert & Ordner umbenannt.\n")
        log("➡️ Jetzt Bilder in die Ring-Ordner legen (Tool offen lassen). Dann START.\n\n")

        try:
            if os.name == "nt":
                os.startfile(project_dir)
        except Exception:
            pass

        show(step4)

    nav3 = tk.Frame(step3)
    nav3.pack(fill="x", padx=10, pady=10)

    ttk.Button(nav3, text="← Zurück", command=step3_back).pack(side="left")

    def step3_next_guard():
        def work():
            step3_save_and_rename()
        try:
            run_busy(work, "Speichere & benenne um…")
        except Exception as e:
            messagebox.showerror("Fehler beim Umbenennen", str(e))

    ttk.Button(nav3, text="Speichern & Umbenennen →", command=step3_next_guard).pack(side="right")

    # STEP 4
    step4 = ttk.LabelFrame(container, text="Schritt 4/4 – Bilder einfügen & Start")

    s4 = tk.Frame(step4)
    s4.pack(fill="x", padx=10, pady=10)

    tk.Label(
        s4,
        text="1) Lass dieses Fenster offen\n2) Pack Screenshots in die Ring-Ordner\n3) Klick START",
        foreground="#444"
    ).pack(anchor="w")

    nav4 = tk.Frame(step4)
    nav4.pack(fill="x", padx=10, pady=10)

    def step4_back():
        run_busy(lambda: show(step3), "Wechsle Schritt…")

    def worker():
        try:
            state["running"] = True

            old_out, old_err = sys.stdout, sys.stderr
            sys.stdout = GuiLogger(qlog)
            sys.stderr = GuiLogger(qlog)

            try:
                run_project(
                    state["project_dir"],
                    state["site"],
                    state["rings_final"],
                    log=lambda s: qlog.put(s),
                )
            finally:
                sys.stdout, sys.stderr = old_out, old_err

        except Exception as e:
            qlog.put(f"\n❌ ERROR: {e}\n")

        finally:
            state["running"] = False
            qlog.put("\nFertig.\n")
            root.after(0, lambda: busy.stop())

    def start_run():
        if state["running"]:
            messagebox.showinfo("Läuft schon", "Das Tool läuft bereits.")
            return
        if not state["rings_final"]:
            messagebox.showwarning("Nicht bereit", "Bitte erst Schritt 1-3 abschließen.")
            return

        busy.start("Verarbeite Bilder…")
        root.update_idletasks()
        threading.Thread(target=worker, daemon=True).start()

    ttk.Button(nav4, text="← Zurück", command=step4_back).pack(side="left")
    ttk.Button(nav4, text="START", command=start_run).pack(side="right")

    busy.stop()
    root.deiconify()   # <-- jetzt erst Hauptfenster zeigen
    root.update_idletasks()
    show(step1)
    root.mainloop()


if __name__ == "__main__":
    start_gui()
